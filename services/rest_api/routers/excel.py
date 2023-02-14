from datetime import datetime
from random import randint, choice, sample
import string
from io import BytesIO
from openpyxl import Workbook
from fastapi import APIRouter, Response


router = APIRouter()

@router.get('/files')
async def table()->Response:
    today = datetime.now()

    random_chars = [choice(string.ascii_letters) for i in range(randint(1, 6))] + [choice(string.digits) for i in range(randint(1, 6))]
    random_str = ''.join(sample(random_chars, len(random_chars)))

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Дата'
    ws['B1'] = 'Время'
    ws['C1'] = 'Случайное число'
    ws['D1'] = 'Случайная строка'

    ws['A2'] = today.date()
    ws['B2'] = today.time()
    ws['C2'] = randint(0, 1000)
    ws['D2'] = random_str

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    file_date = today.date().strftime('%Y%m%d')

    file_name = f'file_generated_at_{file_date}.xlsx'

    return Response(file.read(), headers = {'Content-Disposition': f'attachment; filename={file_name}'})
